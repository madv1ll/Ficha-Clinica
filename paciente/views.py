from django.forms.models import model_to_dict
from django.http import  JsonResponse
from django.http.response import HttpResponse
from django.shortcuts import get_object_or_404, render
from django.views import View

from .utils import render_to_pdf
from .models import Evaluacion, Historial,  Medico, Paciente, SignosVitales
from .forms import EvolucionForm, HistorialForm, MedicoForm, PacienteForm,  SignosForm
from django.shortcuts import redirect
from django.views.generic import ListView, CreateView
from django.urls import reverse_lazy
from django.utils.decorators import method_decorator
from django.views.decorators.csrf import csrf_exempt
from django.contrib import messages
from django.utils import timezone
from django.db.models import Q
from django.views.generic.base import TemplateView
from openpyxl import Workbook
from openpyxl.styles import Alignment,Border,Font,PatternFill,Side

def historial(request, rut):
    pacientes = Paciente.objects.filter(rut = rut)
    historial = Historial.objects.filter(rut = rut)
    datos = {
        'pacientes':pacientes,
        'historial':historial
    }
    return render(request, 'historial.html',datos)

def historialDetalle(request, id):
    historialobj = Historial.objects.filter(idhistorial = id)
    datos = {
        'historial': historialobj
    }
    return render(request, 'historialdetalle.html',datos)

def nuevoPaciente(request):
    if request.method == "POST":
        form = PacienteForm(request.POST)
        if form.is_valid():
            post = form.save(commit = False)
            post.nombreMedico = request.user.username
            post.nombreMedicoAdmin = 'cuidador'
            post.save()
            return redirect ('index')
    else:
        form = PacienteForm
    return render(request, 'nuevopaciente.html', {'form':form})

def nuevoHistorialf(request, rut):
    form = HistorialForm(request.POST)
    if request.method == "POST":
        if form.is_valid():
            post = form.save(commit = False)
            post.rut_id = rut
            post.save()
            return redirect ('historial_clinico',rut)
    else:
        form = HistorialForm
    return render(request, 'HistorialForm.html', {'form':form})

def editarPaciente(request, rut):
    post = get_object_or_404(Paciente, rut=rut)
    if request.method == "POST":
        form = PacienteForm(request.POST, instance=post)
        if form.is_valid():
            form.save()
            return redirect('index')
    else:
        form = PacienteForm(instance=post)
    return render(request, 'editarpaciente.html', {'form': form})

def eliminarPaciente(request,rut):
    paciente = Paciente.objects.filter(rut=rut)
    paciente.delete()
    messages.success(request, 'Paciente elminado con éxito')

    return redirect('index')

def pacienteinicio(request):
    pacientes = Paciente.objects.filter(nombreMedico_id = request.user)
    traspaso = {
        'pacientes':pacientes
    }
    return render(request, 'index.html' ,traspaso)

def signosVitales(request, rut):
    pacientes = Paciente.objects.filter(rut = rut)
    signos = SignosVitales.objects.filter(paciente_rut = rut)
    datos = {
        'pacientes':pacientes,
        'signosVitales': signos
    }
    return render(request, 'signosVitales.html',datos)

def nuevoSignosVitales(request, rut):
    form = SignosForm(request.POST)
    if request.method == "POST":
        if form.is_valid():
            post = form.save(commit = False)
            post.paciente_rut_id = rut
            post.save()
            return redirect ('signosVitales',rut)
    else:
        form = SignosForm
    return render(request, 'signosVitalesForm.html', {'form':form})

def signosDetalle(request, id):
    signos = SignosVitales.objects.filter(id_signosvitales = id)
    datos = {
        'signos': signos
    }
    return render(request, 'signosdetalle.html',datos)

class NuevoMedico(CreateView):
    model = Medico
    form_class = MedicoForm
    template_name = 'nuevomedico.html'
    success_url = reverse_lazy('index')

def listaCuidadores(request):
    usuario = Medico.objects.all()
    datos = {
        'usuario': usuario
    }
    return render(request, 'cuidadores.html',datos)

def editarCuidador(request, rut):
    post = get_object_or_404(Medico, rut = rut)
    if request.method == "POST":
        form = MedicoForm(request.POST, instance=post)
        if form.is_valid():
            form.save()
            return redirect('index')
    else:
        form = MedicoForm(instance=post)
    return render(request, 'nuevomedico.html', {'form': form})

def editarHistorial(request, id):
    post = get_object_or_404(Historial, idhistorial=id)
    if request.method == "POST":
        form = HistorialForm(request.POST, instance=post)
        if form.is_valid():
            form.save()
            return redirect('index')
    else:
        form = HistorialForm(instance=post)
    return render(request, 'editarHistorial.html', {'form': form})

def editarSignos(request, id):
    post = get_object_or_404(SignosVitales, id_signosvitales=id)
    if request.method == "POST":
        form = SignosForm(request.POST, instance=post)
        if form.is_valid():
            form.save()
            return redirect('index')
    else:
        form = SignosForm(instance=post)
    return render(request, 'editarSignos.html', {'form': form})

class Index(CreateView):
    model = Paciente
    fields = ('__all__')
    template_name = 'index.html'
    success_url = '.'

    @method_decorator(csrf_exempt)
    def dispatch(self, request, *args, **kwargs):
        return super().dispatch(request, *args, **kwargs)
    
    def post(self, request, *args, **kwargs):
        queryset = request.POST["buscar"]
        if queryset:
            data = list(Paciente.objects.filter(Q(rut__icontains=queryset) | Q(pnombre__icontains=queryset) | Q(papellido__icontains=queryset) ).distinct().values())
        else:
            data = list(Paciente.objects.filter(lugarAtencion=request.POST['id']).values())
        return JsonResponse({'lugaratencion':data})
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["lugarSeleccion"] = Paciente.objects.filter(lugarAtencion=1)
        return context

def evolucion(request, rut):
    pacientes = Paciente.objects.filter(rut = rut)
    evaluacion = Evaluacion.objects.filter(rut = rut)
    datos = {
        'pacientes':pacientes,
        'evolucion':evaluacion
    }
    return render(request, 'evolucion.html',datos)

def nuevaEvolucion(request, rut):
    form = EvolucionForm(request.POST)
    if request.method == "POST":
        if form.is_valid():
            post = form.save(commit = False)
            post.rut_id = rut
            fecha = timezone.now()
            convertida = fecha.strftime("%d-%m-%Y")
            post.fecha_evaluacion = convertida
            post.hora = fecha.strftime('%H:%M')
            post.save()
            return redirect ('evolucion', rut)
    else:
        form = EvolucionForm
    return render(request, 'evolucionForm.html', {'form':form})

class ReporteExcel(TemplateView):
    def get(self, request, *args, **kwargs):
        rut = self.kwargs['rut']
        query = Paciente.objects.filter(rut = rut)
        historial = Historial.objects.filter(rut = rut)
        signosvitales = SignosVitales.objects.filter(paciente_rut = rut)
        evolucion = Evaluacion.objects.filter(rut = rut)
        wb = Workbook()
        controlador = 4
        for q in query:
            ws = wb.active
            ws.title = 'Paciente'
            ws['B1'].alignment = Alignment(horizontal= "center", vertical= "center")
            ws['B1'].border = Border(left = Side(border_style= "thin"), right = Side(border_style= "thin"),
                                     top  = Side(border_style= "thin"), bottom = Side(border_style= "thin"))
            ws['B1'].fill = PatternFill(start_color= 'e3f2fd', end_color= 'e3f2fd', fill_type= "solid")
            ws['B1'].font = Font(name = 'Calibri', size = 12, bold= True)
            ws['B1'] = 'Datos del paciente'

            #dimesiones de la tabla 
            ws.merge_cells('B1:D1')
            ws.row_dimensions[1].height = 25
            ws.column_dimensions['B'].width = 20
            ws.column_dimensions['C'].width = 20
            ws.column_dimensions['D'].width = 20

            ws['B3'].alignment = Alignment(horizontal= "center", vertical= "center")
            ws['B3'].border = Border(left = Side(border_style= "thin"), right = Side(border_style= "thin"),
                                     top  = Side(border_style= "thin"), bottom = Side(border_style= "thin"))
            ws['B3'].fill = PatternFill(start_color= 'e3f2fd', end_color= 'e3f2fd', fill_type= "solid")
            ws['B3'].font = Font(name = 'Calibri', size = 10, bold= True)
            ws['B3'] = 'Rut'

            ws['C3'].alignment = Alignment(horizontal= "center", vertical= "center")
            ws['C3'].border = Border(left = Side(border_style= "thin"), right = Side(border_style= "thin"),
                                     top  = Side(border_style= "thin"), bottom = Side(border_style= "thin"))
            ws['C3'].fill = PatternFill(start_color= 'e3f2fd', end_color= 'e3f2fd', fill_type= "solid")
            ws['C3'].font = Font(name = 'Calibri', size = 10, bold= True)
            ws['C3'] = 'Nombres'

            ws['D3'].alignment = Alignment(horizontal= "center", vertical= "center")
            ws['D3'].border = Border(left = Side(border_style= "thin"), right = Side(border_style= "thin"),
                                     top  = Side(border_style= "thin"), bottom = Side(border_style= "thin"))
            ws['D3'].fill = PatternFill(start_color= 'e3f2fd', end_color= 'e3f2fd', fill_type= "solid")
            ws['D3'].font = Font(name = 'Calibri', size = 10, bold= True)
            ws['D3'] = 'Apellidos'


            #tabla dinamica con los datos
            ws.cell(row = controlador, column= 2).alignment = Alignment(horizontal = "center")
            ws.cell(row = controlador, column= 2).border = Border(left = Side(border_style= "thin"), right = Side(border_style= "thin"),
                                                                  top  = Side(border_style= "thin"), bottom = Side(border_style= "thin"))
            ws.cell(row = controlador, column= 2).font = Font(name = 'Calibri', size = 10)
            ws.cell(row = controlador, column= 2).value = q.rut

            ws.cell(row = controlador, column= 3).alignment = Alignment(horizontal = "center")
            ws.cell(row = controlador, column= 3).border = Border(left = Side(border_style= "thin"), right = Side(border_style= "thin"),
                                                                  top  = Side(border_style= "thin"), bottom = Side(border_style= "thin"))
            ws.cell(row = controlador, column= 3).font = Font(name = 'Calibri', size = 10)
            ws.cell(row = controlador, column= 3).value = q.pnombre +' '+ q.snombre

            ws.cell(row = controlador, column= 4).alignment = Alignment(horizontal = "center")
            ws.cell(row = controlador, column= 4).border = Border(left = Side(border_style= "thin"), right = Side(border_style= "thin"),
                                                                  top  = Side(border_style= "thin"), bottom = Side(border_style= "thin"))
            ws.cell(row = controlador, column= 4).font = Font(name = 'Calibri', size = 10)
            ws.cell(row = controlador, column= 4).value = q.papellido +' '+ q.sapellido

            #Nueva hoja
            ws = wb.create_sheet('Historial')
            ws['B1'].alignment = Alignment(horizontal= "center", vertical= "center")
            ws['B1'].border = Border(left = Side(border_style= "thin"), right = Side(border_style= "thin"),
                                        top  = Side(border_style= "thin"), bottom = Side(border_style= "thin"))
            ws['B1'].fill = PatternFill(start_color= 'e3f2fd', end_color= 'e3f2fd', fill_type= "solid")
            ws['B1'].font = Font(name = 'Calibri', size = 12, bold= True)
            ws['B1'] = 'Historial del paciente'

            ws.merge_cells('B1:J1')
            ws.row_dimensions[1].height = 25
            ws.column_dimensions['B'].width = 20
            ws.column_dimensions['C'].width = 20
            ws.column_dimensions['D'].width = 20
            ws.column_dimensions['E'].width = 20
            ws.column_dimensions['F'].width = 20
            ws.column_dimensions['G'].width = 20
            ws.column_dimensions['H'].width = 20
            ws.column_dimensions['I'].width = 20
            ws.column_dimensions['J'].width = 20

            ws['B3'].alignment = Alignment(horizontal= "center", vertical= "center")
            ws['B3'].border = Border(left = Side(border_style= "thin"), right = Side(border_style= "thin"),
                                    top  = Side(border_style= "thin"), bottom = Side(border_style= "thin"))
            ws['B3'].fill = PatternFill(start_color= 'e3f2fd', end_color= 'e3f2fd', fill_type= "solid")
            ws['B3'].font = Font(name = 'Calibri', size = 10, bold= True)
            ws['B3'] = 'Fecha'

            ws['C3'].alignment = Alignment(horizontal= "center", vertical= "center")
            ws['C3'].border = Border(left = Side(border_style= "thin"), right = Side(border_style= "thin"),
                                    top  = Side(border_style= "thin"), bottom = Side(border_style= "thin"))
            ws['C3'].fill = PatternFill(start_color= 'e3f2fd', end_color= 'e3f2fd', fill_type= "solid")
            ws['C3'].font = Font(name = 'Calibri', size = 10, bold= True)
            ws['C3'] = 'Servicio'

            ws['D3'].alignment = Alignment(horizontal= "center", vertical= "center")
            ws['D3'].border = Border(left = Side(border_style= "thin"), right = Side(border_style= "thin"),
                                    top  = Side(border_style= "thin"), bottom = Side(border_style= "thin"))
            ws['D3'].fill = PatternFill(start_color= 'e3f2fd', end_color= 'e3f2fd', fill_type= "solid")
            ws['D3'].font = Font(name = 'Calibri', size = 10, bold= True)
            ws['D3'] = 'Diagnostico'

            ws['E3'].alignment = Alignment(horizontal= "center", vertical= "center")
            ws['E3'].border = Border(left = Side(border_style= "thin"), right = Side(border_style= "thin"),
                                    top  = Side(border_style= "thin"), bottom = Side(border_style= "thin"))
            ws['E3'].fill = PatternFill(start_color= 'e3f2fd', end_color= 'e3f2fd', fill_type= "solid")
            ws['E3'].font = Font(name = 'Calibri', size = 10, bold= True)
            ws['E3'] = 'Motivo de ingreso' 

            ws['F3'].alignment = Alignment(horizontal= "center", vertical= "center")
            ws['F3'].border = Border(left = Side(border_style= "thin"), right = Side(border_style= "thin"),
                                    top  = Side(border_style= "thin"), bottom = Side(border_style= "thin"))
            ws['F3'].fill = PatternFill(start_color= 'e3f2fd', end_color= 'e3f2fd', fill_type= "solid")
            ws['F3'].font = Font(name = 'Calibri', size = 10, bold= True)
            ws['F3'] = 'Enfermedad actual'

            ws['G3'].alignment = Alignment(horizontal= "center", vertical= "center")
            ws['G3'].border = Border(left = Side(border_style= "thin"), right = Side(border_style= "thin"),
                                    top  = Side(border_style= "thin"), bottom = Side(border_style= "thin"))
            ws['G3'].fill = PatternFill(start_color= 'e3f2fd', end_color= 'e3f2fd', fill_type= "solid")
            ws['G3'].font = Font(name = 'Calibri', size = 10, bold= True)
            ws['G3'] = 'Diagnostico de admision' 

            ws['H3'].alignment = Alignment(horizontal= "center", vertical= "center")
            ws['H3'].border = Border(left = Side(border_style= "thin"), right = Side(border_style= "thin"),
                                    top  = Side(border_style= "thin"), bottom = Side(border_style= "thin"))
            ws['H3'].fill = PatternFill(start_color= 'e3f2fd', end_color= 'e3f2fd', fill_type= "solid")
            ws['H3'].font = Font(name = 'Calibri', size = 10, bold= True)
            ws['H3'] = 'Diagnostico clinico final'  

            ws['I3'].alignment = Alignment(horizontal= "center", vertical= "center")
            ws['I3'].border = Border(left = Side(border_style= "thin"), right = Side(border_style= "thin"),
                                    top  = Side(border_style= "thin"), bottom = Side(border_style= "thin"))
            ws['I3'].fill = PatternFill(start_color= 'e3f2fd', end_color= 'e3f2fd', fill_type= "solid")
            ws['I3'].font = Font(name = 'Calibri', size = 10, bold= True)
            ws['I3'] = 'Fecha alta medica'  

            ws['J3'].alignment = Alignment(horizontal= "center", vertical= "center")
            ws['J3'].border = Border(left = Side(border_style= "thin"), right = Side(border_style= "thin"),
                                    top  = Side(border_style= "thin"), bottom = Side(border_style= "thin"))
            ws['J3'].fill = PatternFill(start_color= 'e3f2fd', end_color= 'e3f2fd', fill_type= "solid")
            ws['J3'].font = Font(name = 'Calibri', size = 10, bold= True)
            ws['J3'] = 'Fecha alta clinica'   


            for hist in historial: 
                if hist.fecha_alta_clinica != None:
                    fecha_alta_clinica = hist.fecha_alta_clinica.strftime('%d/%m/%Y')
                else:
                    fecha_alta_clinica = ""
                if hist.fecha_alta_medica != None:
                    fecha_alta_medica = hist.fecha_alta_medica.strftime('%d/%m/%Y')
                else:
                    fecha_alta_medica = ""

                #tabla dinamica con los datos
                ws.cell(row = controlador, column= 2).alignment = Alignment(horizontal = "center")
                ws.cell(row = controlador, column= 2).border = Border(left = Side(border_style= "thin"), right = Side(border_style= "thin"),
                                                                    top  = Side(border_style= "thin"), bottom = Side(border_style= "thin"))
                ws.cell(row = controlador, column= 2).font = Font(name = 'Calibri', size = 10)
                ws.cell(row = controlador, column= 2).value = hist.fecha.strftime('%d/%m/%Y')                

                ws.cell(row = controlador, column= 3).alignment = Alignment(horizontal = "center")
                ws.cell(row = controlador, column= 3).border = Border(left = Side(border_style= "thin"), right = Side(border_style= "thin"),
                                                                    top  = Side(border_style= "thin"), bottom = Side(border_style= "thin"))
                ws.cell(row = controlador, column= 3).font = Font(name = 'Calibri', size = 10)
                ws.cell(row = controlador, column= 3).value = hist.servicio

                ws.cell(row = controlador, column= 4).alignment = Alignment(horizontal = "center")
                ws.cell(row = controlador, column= 4).border = Border(left = Side(border_style= "thin"), right = Side(border_style= "thin"),
                                                                    top  = Side(border_style= "thin"), bottom = Side(border_style= "thin"))
                ws.cell(row = controlador, column= 4).font = Font(name = 'Calibri', size = 10)
                ws.cell(row = controlador, column= 4).value = hist.diagnostico

                ws.cell(row = controlador, column= 5).alignment = Alignment(horizontal = "center")
                ws.cell(row = controlador, column= 5).border = Border(left = Side(border_style= "thin"), right = Side(border_style= "thin"),
                                                                    top  = Side(border_style= "thin"), bottom = Side(border_style= "thin"))
                ws.cell(row = controlador, column= 5).font = Font(name = 'Calibri', size = 10)
                ws.cell(row = controlador, column= 5).value = hist.motivo_ingreso

                ws.cell(row = controlador, column= 6).alignment = Alignment(horizontal = "center")
                ws.cell(row = controlador, column= 6).border = Border(left = Side(border_style= "thin"), right = Side(border_style= "thin"),
                                                                    top  = Side(border_style= "thin"), bottom = Side(border_style= "thin"))
                ws.cell(row = controlador, column= 6).font = Font(name = 'Calibri', size = 10)
                ws.cell(row = controlador, column= 6).value = hist.enfermedad_actual

                ws.cell(row = controlador, column= 7).alignment = Alignment(horizontal = "center")
                ws.cell(row = controlador, column= 7).border = Border(left = Side(border_style= "thin"), right = Side(border_style= "thin"),
                                                                    top  = Side(border_style= "thin"), bottom = Side(border_style= "thin"))
                ws.cell(row = controlador, column= 7).font = Font(name = 'Calibri', size = 10)
                ws.cell(row = controlador, column= 7).value = hist.diagnostico_admision

                ws.cell(row = controlador, column= 8).alignment = Alignment(horizontal = "center")
                ws.cell(row = controlador, column= 8).border = Border(left = Side(border_style= "thin"), right = Side(border_style= "thin"),
                                                                    top  = Side(border_style= "thin"), bottom = Side(border_style= "thin"))
                ws.cell(row = controlador, column= 8).font = Font(name = 'Calibri', size = 10)
                ws.cell(row = controlador, column= 8).value = hist.diagnostico_clinico_final

                ws.cell(row = controlador, column= 9).alignment = Alignment(horizontal = "center")
                ws.cell(row = controlador, column= 9).border = Border(left = Side(border_style= "thin"), right = Side(border_style= "thin"),
                                                                    top  = Side(border_style= "thin"), bottom = Side(border_style= "thin"))
                ws.cell(row = controlador, column= 9).font = Font(name = 'Calibri', size = 10)
                ws.cell(row = controlador, column= 9).value = fecha_alta_medica

                ws.cell(row = controlador, column= 10).alignment = Alignment(horizontal = "center")
                ws.cell(row = controlador, column= 10).border = Border(left = Side(border_style= "thin"), right = Side(border_style= "thin"),
                                                                    top  = Side(border_style= "thin"), bottom = Side(border_style= "thin"))
                ws.cell(row = controlador, column= 10).font = Font(name = 'Calibri', size = 10)
                ws.cell(row = controlador, column= 10).value = fecha_alta_clinica
                controlador+=1


            ws = wb.create_sheet('Signos vitales')
            ws['B1'].alignment = Alignment(horizontal= "center", vertical= "center")
            ws['B1'].border = Border(left = Side(border_style= "thin"), right = Side(border_style= "thin"),
                                        top  = Side(border_style= "thin"), bottom = Side(border_style= "thin"))
            ws['B1'].fill = PatternFill(start_color= 'e3f2fd', end_color= 'e3f2fd', fill_type= "solid")
            ws['B1'].font = Font(name = 'Calibri', size = 12, bold= True)
            ws['B1'] = 'Signos del paciente'

            ws.merge_cells('B1:T1')
            ws.row_dimensions[1].height = 25
            ws.column_dimensions['B'].width = 20
            ws.column_dimensions['C'].width = 20
            ws.column_dimensions['D'].width = 20
            ws.column_dimensions['E'].width = 20
            ws.column_dimensions['F'].width = 20
            ws.column_dimensions['G'].width = 20
            ws.column_dimensions['H'].width = 20
            ws.column_dimensions['I'].width = 20
            ws.column_dimensions['J'].width = 20
            ws.column_dimensions['K'].width = 20
            ws.column_dimensions['L'].width = 20
            ws.column_dimensions['M'].width = 20
            ws.column_dimensions['N'].width = 20
            ws.column_dimensions['O'].width = 20
            ws.column_dimensions['P'].width = 20
            ws.column_dimensions['Q'].width = 20
            ws.column_dimensions['R'].width = 20
            ws.column_dimensions['S'].width = 20
            ws.column_dimensions['T'].width = 20
            controlador = 4


            ws['B3'].alignment = Alignment(horizontal= "center", vertical= "center")
            ws['B3'].border = Border(left = Side(border_style= "thin"), right = Side(border_style= "thin"),
                                    top  = Side(border_style= "thin"), bottom = Side(border_style= "thin"))
            ws['B3'].fill = PatternFill(start_color= 'e3f2fd', end_color= 'e3f2fd', fill_type= "solid")
            ws['B3'].font = Font(name = 'Calibri', size = 10, bold= True)
            ws['B3'] = 'temperatura'

            ws['C3'].alignment = Alignment(horizontal= "center", vertical= "center")
            ws['C3'].border = Border(left = Side(border_style= "thin"), right = Side(border_style= "thin"),
                                    top  = Side(border_style= "thin"), bottom = Side(border_style= "thin"))
            ws['C3'].fill = PatternFill(start_color= 'e3f2fd', end_color= 'e3f2fd', fill_type= "solid")
            ws['C3'].font = Font(name = 'Calibri', size = 10, bold= True)
            ws['C3'] = 'respiracion'

            ws['D3'].alignment = Alignment(horizontal= "center", vertical= "center")
            ws['D3'].border = Border(left = Side(border_style= "thin"), right = Side(border_style= "thin"),
                                    top  = Side(border_style= "thin"), bottom = Side(border_style= "thin"))
            ws['D3'].fill = PatternFill(start_color= 'e3f2fd', end_color= 'e3f2fd', fill_type= "solid")
            ws['D3'].font = Font(name = 'Calibri', size = 10, bold= True)
            ws['D3'] = 'presion arterial' 

            ws['E3'].alignment = Alignment(horizontal= "center", vertical= "center")
            ws['E3'].border = Border(left = Side(border_style= "thin"), right = Side(border_style= "thin"),
                                    top  = Side(border_style= "thin"), bottom = Side(border_style= "thin"))
            ws['E3'].fill = PatternFill(start_color= 'e3f2fd', end_color= 'e3f2fd', fill_type= "solid")
            ws['E3'].font = Font(name = 'Calibri', size = 10, bold= True)
            ws['E3'] = 'evaluacion'

            ws['F3'].alignment = Alignment(horizontal= "center", vertical= "center")
            ws['F3'].border = Border(left = Side(border_style= "thin"), right = Side(border_style= "thin"),
                                    top  = Side(border_style= "thin"), bottom = Side(border_style= "thin"))
            ws['F3'].fill = PatternFill(start_color= 'e3f2fd', end_color= 'e3f2fd', fill_type= "solid")
            ws['F3'].font = Font(name = 'Calibri', size = 10, bold= True)
            ws['F3'] = 'miccion' 

            ws['G3'].alignment = Alignment(horizontal= "center", vertical= "center")
            ws['G3'].border = Border(left = Side(border_style= "thin"), right = Side(border_style= "thin"),
                                    top  = Side(border_style= "thin"), bottom = Side(border_style= "thin"))
            ws['G3'].fill = PatternFill(start_color= 'e3f2fd', end_color= 'e3f2fd', fill_type= "solid")
            ws['G3'].font = Font(name = 'Calibri', size = 10, bold= True)
            ws['G3'] = 'vomito'

            ws['H3'].alignment = Alignment(horizontal= "center", vertical= "center")
            ws['H3'].border = Border(left = Side(border_style= "thin"), right = Side(border_style= "thin"),
                                    top  = Side(border_style= "thin"), bottom = Side(border_style= "thin"))
            ws['H3'].fill = PatternFill(start_color= 'e3f2fd', end_color= 'e3f2fd', fill_type= "solid")
            ws['H3'].font = Font(name = 'Calibri', size = 10, bold= True)
            ws['H3'] = 'flatos'

            ws['I3'].alignment = Alignment(horizontal= "center", vertical= "center")
            ws['I3'].border = Border(left = Side(border_style= "thin"), right = Side(border_style= "thin"),
                                    top  = Side(border_style= "thin"), bottom = Side(border_style= "thin"))
            ws['I3'].fill = PatternFill(start_color= 'e3f2fd', end_color= 'e3f2fd', fill_type= "solid")
            ws['I3'].font = Font(name = 'Calibri', size = 10, bold= True)
            ws['I3'] = 'dolor '

            ws['J3'].alignment = Alignment(horizontal= "center", vertical= "center")
            ws['J3'].border = Border(left = Side(border_style= "thin"), right = Side(border_style= "thin"),
                                    top  = Side(border_style= "thin"), bottom = Side(border_style= "thin"))
            ws['J3'].fill = PatternFill(start_color= 'e3f2fd', end_color= 'e3f2fd', fill_type= "solid")
            ws['J3'].font = Font(name = 'Calibri', size = 10, bold= True)
            ws['J3'] = 'estrennimiento ' 

            ws['K3'].alignment = Alignment(horizontal= "center", vertical= "center")
            ws['K3'].border = Border(left = Side(border_style= "thin"), right = Side(border_style= "thin"),
                                    top  = Side(border_style= "thin"), bottom = Side(border_style= "thin"))
            ws['K3'].fill = PatternFill(start_color= 'e3f2fd', end_color= 'e3f2fd', fill_type= "solid")
            ws['K3'].font = Font(name = 'Calibri', size = 10, bold= True)
            ws['K3'] = 'frecuencia cardiaca  '

            ws['L3'].alignment = Alignment(horizontal= "center", vertical= "center")
            ws['L3'].border = Border(left = Side(border_style= "thin"), right = Side(border_style= "thin"),
                                    top  = Side(border_style= "thin"), bottom = Side(border_style= "thin"))
            ws['L3'].fill = PatternFill(start_color= 'e3f2fd', end_color= 'e3f2fd', fill_type= "solid")
            ws['L3'].font = Font(name = 'Calibri', size = 10, bold= True)
            ws['L3'] = 'saturacion ' 

            ws['M3'].alignment = Alignment(horizontal= "center", vertical= "center")
            ws['M3'].border = Border(left = Side(border_style= "thin"), right = Side(border_style= "thin"),
                                    top  = Side(border_style= "thin"), bottom = Side(border_style= "thin"))
            ws['M3'].fill = PatternFill(start_color= 'e3f2fd', end_color= 'e3f2fd', fill_type= "solid")
            ws['M3'].font = Font(name = 'Calibri', size = 10, bold= True)
            ws['M3'] = 'suenno '  

            ws['N3'].alignment = Alignment(horizontal= "center", vertical= "center")
            ws['N3'].border = Border(left = Side(border_style= "thin"), right = Side(border_style= "thin"),
                                    top  = Side(border_style= "thin"), bottom = Side(border_style= "thin"))
            ws['N3'].fill = PatternFill(start_color= 'e3f2fd', end_color= 'e3f2fd', fill_type= "solid")
            ws['N3'].font = Font(name = 'Calibri', size = 10, bold= True)
            ws['N3'] = 'presion arterial media'

            ws['O3'].alignment = Alignment(horizontal= "center", vertical= "center")
            ws['O3'].border = Border(left = Side(border_style= "thin"), right = Side(border_style= "thin"),
                                    top  = Side(border_style= "thin"), bottom = Side(border_style= "thin"))
            ws['O3'].fill = PatternFill(start_color= 'e3f2fd', end_color= 'e3f2fd', fill_type= "solid")
            ws['O3'].font = Font(name = 'Calibri', size = 10, bold= True)
            ws['O3'] = 'peso'

            ws['P3'].alignment = Alignment(horizontal= "center", vertical= "center")
            ws['P3'].border = Border(left = Side(border_style= "thin"), right = Side(border_style= "thin"),
                                    top  = Side(border_style= "thin"), bottom = Side(border_style= "thin"))
            ws['P3'].fill = PatternFill(start_color= 'e3f2fd', end_color= 'e3f2fd', fill_type= "solid")
            ws['P3'].font = Font(name = 'Calibri', size = 10, bold= True)
            ws['P3'] = 'talla' 

            ws['Q3'].alignment = Alignment(horizontal= "center", vertical= "center")
            ws['Q3'].border = Border(left = Side(border_style= "thin"), right = Side(border_style= "thin"),
                                    top  = Side(border_style= "thin"), bottom = Side(border_style= "thin"))
            ws['Q3'].fill = PatternFill(start_color= 'e3f2fd', end_color= 'e3f2fd', fill_type= "solid")
            ws['Q3'].font = Font(name = 'Calibri', size = 10, bold= True)
            ws['Q3'] = 'alimentacion'

            ws['R3'].alignment = Alignment(horizontal= "center", vertical= "center")
            ws['R3'].border = Border(left = Side(border_style= "thin"), right = Side(border_style= "thin"),
                                    top  = Side(border_style= "thin"), bottom = Side(border_style= "thin"))
            ws['R3'].fill = PatternFill(start_color= 'e3f2fd', end_color= 'e3f2fd', fill_type= "solid")
            ws['R3'].font = Font(name = 'Calibri', size = 10, bold= True)
            ws['R3'] = 'higiene' 

            ws['S3'].alignment = Alignment(horizontal= "center", vertical= "center")
            ws['S3'].border = Border(left = Side(border_style= "thin"), right = Side(border_style= "thin"),
                                    top  = Side(border_style= "thin"), bottom = Side(border_style= "thin"))
            ws['S3'].fill = PatternFill(start_color= 'e3f2fd', end_color= 'e3f2fd', fill_type= "solid")
            ws['S3'].font = Font(name = 'Calibri', size = 10, bold= True)
            ws['S3'] = 'observaciones'  

            ws['T3'].alignment = Alignment(horizontal= "center", vertical= "center")
            ws['T3'].border = Border(left = Side(border_style= "thin"), right = Side(border_style= "thin"),
                                    top  = Side(border_style= "thin"), bottom = Side(border_style= "thin"))
            ws['T3'].fill = PatternFill(start_color= 'e3f2fd', end_color= 'e3f2fd', fill_type= "solid")
            ws['T3'].font = Font(name = 'Calibri', size = 10, bold= True)
            ws['T3'] = 'Fecha de creacion' 

            for sig in signosvitales:
                #tabla dinamica con los datos

                ws.cell(row = controlador, column= 2).alignment = Alignment(horizontal = "center")
                ws.cell(row = controlador, column= 2).border = Border(left = Side(border_style= "thin"), right = Side(border_style= "thin"),
                                                                    top  = Side(border_style= "thin"), bottom = Side(border_style= "thin"))
                ws.cell(row = controlador, column= 2).font = Font(name = 'Calibri', size = 10)
                ws.cell(row = controlador, column= 2).value = sig.temperatura

                ws.cell(row = controlador, column= 3).alignment = Alignment(horizontal = "center")
                ws.cell(row = controlador, column= 3).border = Border(left = Side(border_style= "thin"), right = Side(border_style= "thin"),
                                                                    top  = Side(border_style= "thin"), bottom = Side(border_style= "thin"))
                ws.cell(row = controlador, column= 3).font = Font(name = 'Calibri', size = 10)
                ws.cell(row = controlador, column= 3).value = sig.respiracion

                ws.cell(row = controlador, column= 4).alignment = Alignment(horizontal = "center")
                ws.cell(row = controlador, column= 4).border = Border(left = Side(border_style= "thin"), right = Side(border_style= "thin"),
                                                                    top  = Side(border_style= "thin"), bottom = Side(border_style= "thin"))
                ws.cell(row = controlador, column= 4).font = Font(name = 'Calibri', size = 10)
                ws.cell(row = controlador, column= 4).value = sig.presion_arterial

                ws.cell(row = controlador, column= 5).alignment = Alignment(horizontal = "center")
                ws.cell(row = controlador, column= 5).border = Border(left = Side(border_style= "thin"), right = Side(border_style= "thin"),
                                                                    top  = Side(border_style= "thin"), bottom = Side(border_style= "thin"))
                ws.cell(row = controlador, column= 5).font = Font(name = 'Calibri', size = 10)
                ws.cell(row = controlador, column= 5).value = sig.evaluacion

                ws.cell(row = controlador, column= 6).alignment = Alignment(horizontal = "center")
                ws.cell(row = controlador, column= 6).border = Border(left = Side(border_style= "thin"), right = Side(border_style= "thin"),
                                                                    top  = Side(border_style= "thin"), bottom = Side(border_style= "thin"))
                ws.cell(row = controlador, column= 6).font = Font(name = 'Calibri', size = 10)
                ws.cell(row = controlador, column= 6).value = sig.miccion

                ws.cell(row = controlador, column= 7).alignment = Alignment(horizontal = "center")
                ws.cell(row = controlador, column= 7).border = Border(left = Side(border_style= "thin"), right = Side(border_style= "thin"),
                                                                    top  = Side(border_style= "thin"), bottom = Side(border_style= "thin"))
                ws.cell(row = controlador, column= 7).font = Font(name = 'Calibri', size = 10)
                ws.cell(row = controlador, column= 7).value = sig.vomito

                ws.cell(row = controlador, column= 8).alignment = Alignment(horizontal = "center")
                ws.cell(row = controlador, column= 8).border = Border(left = Side(border_style= "thin"), right = Side(border_style= "thin"),
                                                                    top  = Side(border_style= "thin"), bottom = Side(border_style= "thin"))
                ws.cell(row = controlador, column= 8).font = Font(name = 'Calibri', size = 10)
                ws.cell(row = controlador, column= 8).value = sig.flatos

                ws.cell(row = controlador, column= 9).alignment = Alignment(horizontal = "center")
                ws.cell(row = controlador, column= 9).border = Border(left = Side(border_style= "thin"), right = Side(border_style= "thin"),
                                                                    top  = Side(border_style= "thin"), bottom = Side(border_style= "thin"))
                ws.cell(row = controlador, column= 9).font = Font(name = 'Calibri', size = 10)
                ws.cell(row = controlador, column= 9).value = sig.dolor

                ws.cell(row = controlador, column= 10).alignment = Alignment(horizontal = "center")
                ws.cell(row = controlador, column= 10).border = Border(left = Side(border_style= "thin"), right = Side(border_style= "thin"),
                                                                    top  = Side(border_style= "thin"), bottom = Side(border_style= "thin"))
                ws.cell(row = controlador, column= 10).font = Font(name = 'Calibri', size = 10)
                ws.cell(row = controlador, column= 10).value = sig.estrennimiento

                ws.cell(row = controlador, column= 11).alignment = Alignment(horizontal = "center")
                ws.cell(row = controlador, column= 11).border = Border(left = Side(border_style= "thin"), right = Side(border_style= "thin"),
                                                                    top  = Side(border_style= "thin"), bottom = Side(border_style= "thin"))
                ws.cell(row = controlador, column= 11).font = Font(name = 'Calibri', size = 10)
                ws.cell(row = controlador, column= 11).value = sig.frecuencia_cardiaca

                ws.cell(row = controlador, column= 12).alignment = Alignment(horizontal = "center")
                ws.cell(row = controlador, column= 12).border = Border(left = Side(border_style= "thin"), right = Side(border_style= "thin"),
                                                                    top  = Side(border_style= "thin"), bottom = Side(border_style= "thin"))
                ws.cell(row = controlador, column= 12).font = Font(name = 'Calibri', size = 10)
                ws.cell(row = controlador, column= 12).value = sig.saturacion

                ws.cell(row = controlador, column= 13).alignment = Alignment(horizontal = "center")
                ws.cell(row = controlador, column= 13).border = Border(left = Side(border_style= "thin"), right = Side(border_style= "thin"),
                                                                    top  = Side(border_style= "thin"), bottom = Side(border_style= "thin"))
                ws.cell(row = controlador, column= 13).font = Font(name = 'Calibri', size = 10)
                ws.cell(row = controlador, column= 13).value = sig.suenno

                ws.cell(row = controlador, column= 14).alignment = Alignment(horizontal = "center")
                ws.cell(row = controlador, column= 14).border = Border(left = Side(border_style= "thin"), right = Side(border_style= "thin"),
                                                                    top  = Side(border_style= "thin"), bottom = Side(border_style= "thin"))
                ws.cell(row = controlador, column= 14).font = Font(name = 'Calibri', size = 10)
                ws.cell(row = controlador, column= 14).value = sig.presion_arterial_media

                ws.cell(row = controlador, column= 15).alignment = Alignment(horizontal = "center")
                ws.cell(row = controlador, column= 15).border = Border(left = Side(border_style= "thin"), right = Side(border_style= "thin"),
                                                                    top  = Side(border_style= "thin"), bottom = Side(border_style= "thin"))
                ws.cell(row = controlador, column= 15).font = Font(name = 'Calibri', size = 10)
                ws.cell(row = controlador, column= 15).value = sig.peso

                ws.cell(row = controlador, column= 16).alignment = Alignment(horizontal = "center")
                ws.cell(row = controlador, column= 16).border = Border(left = Side(border_style= "thin"), right = Side(border_style= "thin"),
                                                                    top  = Side(border_style= "thin"), bottom = Side(border_style= "thin"))
                ws.cell(row = controlador, column= 16).font = Font(name = 'Calibri', size = 10)
                ws.cell(row = controlador, column= 16).value = sig.talla

                ws.cell(row = controlador, column= 17).alignment = Alignment(horizontal = "center")
                ws.cell(row = controlador, column= 17).border = Border(left = Side(border_style= "thin"), right = Side(border_style= "thin"),
                                                                    top  = Side(border_style= "thin"), bottom = Side(border_style= "thin"))
                ws.cell(row = controlador, column= 17).font = Font(name = 'Calibri', size = 10)
                ws.cell(row = controlador, column= 17).value = sig.alimentacion

                ws.cell(row = controlador, column= 18).alignment = Alignment(horizontal = "center")
                ws.cell(row = controlador, column= 18).border = Border(left = Side(border_style= "thin"), right = Side(border_style= "thin"),
                                                                    top  = Side(border_style= "thin"), bottom = Side(border_style= "thin"))
                ws.cell(row = controlador, column= 18).font = Font(name = 'Calibri', size = 10)
                ws.cell(row = controlador, column= 18).value = sig.higiene

                ws.cell(row = controlador, column= 19).alignment = Alignment(horizontal = "center")
                ws.cell(row = controlador, column= 19).border = Border(left = Side(border_style= "thin"), right = Side(border_style= "thin"),
                                                                    top  = Side(border_style= "thin"), bottom = Side(border_style= "thin"))
                ws.cell(row = controlador, column= 19).font = Font(name = 'Calibri', size = 10)
                ws.cell(row = controlador, column= 19).value = sig.observaciones

                ws.cell(row = controlador, column= 20).alignment = Alignment(horizontal = "center")
                ws.cell(row = controlador, column= 20).border = Border(left = Side(border_style= "thin"), right = Side(border_style= "thin"),
                                                                    top  = Side(border_style= "thin"), bottom = Side(border_style= "thin"))
                ws.cell(row = controlador, column= 20).font = Font(name = 'Calibri', size = 10)
                ws.cell(row = controlador, column= 20).value = sig.fecha_creacion.strftime('%d/%m/%Y')
                controlador+=1

            ws = wb.create_sheet('Evolucion')
            ws['B1'].alignment = Alignment(horizontal= "center", vertical= "center")
            ws['B1'].border = Border(left = Side(border_style= "thin"), right = Side(border_style= "thin"),
                                        top  = Side(border_style= "thin"), bottom = Side(border_style= "thin"))
            ws['B1'].fill = PatternFill(start_color= 'e3f2fd', end_color= 'e3f2fd', fill_type= "solid")
            ws['B1'].font = Font(name = 'Calibri', size = 12, bold= True)
            ws['B1'] = 'Evolucion del cuidador'

            ws.merge_cells('B1:E1')
            ws.row_dimensions[1].height = 25
            ws.column_dimensions['B'].width = 20
            ws.column_dimensions['C'].width = 20
            ws.column_dimensions['D'].width = 20
            ws.column_dimensions['E'].width = 20
            controlador = 4

            ws['B3'].alignment = Alignment(horizontal= "center", vertical= "center")
            ws['B3'].border = Border(left = Side(border_style= "thin"), right = Side(border_style= "thin"),
                                    top  = Side(border_style= "thin"), bottom = Side(border_style= "thin"))
            ws['B3'].fill = PatternFill(start_color= 'e3f2fd', end_color= 'e3f2fd', fill_type= "solid")
            ws['B3'].font = Font(name = 'Calibri', size = 10, bold= True)
            ws['B3'] = 'fecha evaluacion'

            ws['C3'].alignment = Alignment(horizontal= "center", vertical= "center")
            ws['C3'].border = Border(left = Side(border_style= "thin"), right = Side(border_style= "thin"),
                                    top  = Side(border_style= "thin"), bottom = Side(border_style= "thin"))
            ws['C3'].fill = PatternFill(start_color= 'e3f2fd', end_color= 'e3f2fd', fill_type= "solid")
            ws['C3'].font = Font(name = 'Calibri', size = 10, bold= True)
            ws['C3'] = 'hora'

            ws['D3'].alignment = Alignment(horizontal= "center", vertical= "center")
            ws['D3'].border = Border(left = Side(border_style= "thin"), right = Side(border_style= "thin"),
                                    top  = Side(border_style= "thin"), bottom = Side(border_style= "thin"))
            ws['D3'].fill = PatternFill(start_color= 'e3f2fd', end_color= 'e3f2fd', fill_type= "solid")
            ws['D3'].font = Font(name = 'Calibri', size = 10, bold= True)
            ws['D3'] = 'cuidador' 

            ws['E3'].alignment = Alignment(horizontal= "center", vertical= "center")
            ws['E3'].border = Border(left = Side(border_style= "thin"), right = Side(border_style= "thin"),
                                    top  = Side(border_style= "thin"), bottom = Side(border_style= "thin"))
            ws['E3'].fill = PatternFill(start_color= 'e3f2fd', end_color= 'e3f2fd', fill_type= "solid")
            ws['E3'].font = Font(name = 'Calibri', size = 10, bold= True)
            ws['E3'] = 'descripcion'

            for ev in evolucion:

                ws.cell(row = controlador, column= 2).alignment = Alignment(horizontal = "center")
                ws.cell(row = controlador, column= 2).border = Border(left = Side(border_style= "thin"), right = Side(border_style= "thin"),
                                                                    top  = Side(border_style= "thin"), bottom = Side(border_style= "thin"))
                ws.cell(row = controlador, column= 2).font = Font(name = 'Calibri', size = 10)
                ws.cell(row = controlador, column= 2).value = ev.fecha_evaluacion.strftime('%d/%m/%Y')

                ws.cell(row = controlador, column= 3).alignment = Alignment(horizontal = "center")
                ws.cell(row = controlador, column= 3).border = Border(left = Side(border_style= "thin"), right = Side(border_style= "thin"),
                                                                    top  = Side(border_style= "thin"), bottom = Side(border_style= "thin"))
                ws.cell(row = controlador, column= 3).font = Font(name = 'Calibri', size = 10)
                ws.cell(row = controlador, column= 3).value = ev.hora.strftime('%H:%M:%S')

                ws.cell(row = controlador, column= 4).alignment = Alignment(horizontal = "center")
                ws.cell(row = controlador, column= 4).border = Border(left = Side(border_style= "thin"), right = Side(border_style= "thin"),
                                                                    top  = Side(border_style= "thin"), bottom = Side(border_style= "thin"))
                ws.cell(row = controlador, column= 4).font = Font(name = 'Calibri', size = 10)
                ws.cell(row = controlador, column= 4).value = ev.cuidador

                ws.cell(row = controlador, column= 5).alignment = Alignment(horizontal = "center")
                ws.cell(row = controlador, column= 5).border = Border(left = Side(border_style= "thin"), right = Side(border_style= "thin"),
                                                                    top  = Side(border_style= "thin"), bottom = Side(border_style= "thin"))
                ws.cell(row = controlador, column= 5).font = Font(name = 'Calibri', size = 10)
                ws.cell(row = controlador, column= 5).value = ev.descripcion
                controlador+=1



        nombre_archivo = "ReporteExcel.xlsx"
        response = HttpResponse(content_type = "application/ms-excel")
        contenido = "attachment; filename = {0}".format(nombre_archivo)
        response["Content-Disposition"] = contenido
        wb.save(response)
        return response

class ReportePDF(View):
    def get(self, request, *args, **kwargs):
        rut           = self.kwargs['rut']
        template_name = 'reportePDF.html'
        paciente      = Paciente.objects.filter(rut = rut)
        historial     = Historial.objects.filter( rut = rut)
        signos        = SignosVitales.objects.filter( paciente_rut = rut)
        evolucion     = Evaluacion.objects.filter(rut = rut)
        data = {
            'paciente' : paciente,
            'historial': historial,
            'signos'   : signos,
            'evolucion': evolucion
        }
        pdf = render_to_pdf(template_name, data)
        return HttpResponse(pdf, content_type='application/pdf')